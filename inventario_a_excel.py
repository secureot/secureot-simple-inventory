#!/usr/bin/env python3
"""
Convierte la salida del inventario de red (el CSV o JSON que produce network_inventory)
en un Excel con colores, filtros y hojas de resumen.

Uso:
    python3 inventario_a_excel.py inventario.csv               -> inventario.xlsx
    python3 inventario_a_excel.py inventario.json -o reporte.xlsx
"""
import argparse
import csv
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Paleta
AZUL   = "1F4E78"   # encabezados
AZUL2  = "2E75B6"
GRIS   = "D9D9D9"   # filas multicast
NARANJA= "FCE4D6"   # vendor desconocido
AMAR   = "FFF2CC"   # MAC aleatorizada
VERDE  = "C6EFCE"
ROJO   = "FFC7CE"
BLANCO = "FFFFFF"

FONT = "Arial"
thin = Side(style="thin", color="BFBFBF")
BORDE = Border(left=thin, right=thin, top=thin, bottom=thin)


def leer_inventario(path):
    """Devuelve una lista de dicts normalizados desde CSV o JSON."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        with open(path, encoding="utf-8") as f:
            raw = json.load(f)
    else:
        with open(path, encoding="utf-8", newline="") as f:
            raw = list(csv.DictReader(f))

    def b(v):  # normaliza booleanos venidos de CSV ("true"/"false") o JSON (bool)
        return v is True or str(v).strip().lower() in ("true", "1", "si", "sí", "yes")

    filas = []
    for r in raw:
        filas.append({
            "mac": r.get("mac", ""),
            "vendor": r.get("vendor", "") or "Unknown",
            "src": int(r.get("src_count", 0) or 0),
            "dst": int(r.get("dst_count", 0) or 0),
            "multicast": b(r.get("multicast", False)),
            "local": b(r.get("locally_administered", False)),
        })
    return filas


def hoja_inventario(wb, filas):
    ws = wb.active
    ws.title = "Inventario"
    cabeceras = ["MAC", "Fabricante", "Paq. origen", "Paq. destino",
                 "Total", "Multicast", "MAC aleatorizada"]
    n = len(filas)
    ncols = len(cabeceras)
    last = 2 + n  # título=fila1, cabecera=fila2, datos desde fila3

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, "Inventario de dispositivos de red")
    t.font = Font(name=FONT, bold=True, size=14, color=BLANCO)
    t.alignment = Alignment(horizontal="left", vertical="center")
    t.fill = PatternFill("solid", fgColor=AZUL)
    ws.row_dimensions[1].height = 24

    # Cabecera
    for c, txt in enumerate(cabeceras, start=1):
        cell = ws.cell(2, c, txt)
        cell.font = Font(name=FONT, bold=True, color=BLANCO)
        cell.fill = PatternFill("solid", fgColor=AZUL2)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDE

    # Datos (ordenados por total descendente)
    filas = sorted(filas, key=lambda r: r["src"] + r["dst"], reverse=True)
    for i, r in enumerate(filas):
        row = 3 + i
        ws.cell(row, 1, r["mac"]).font = Font(name=FONT)
        ws.cell(row, 2, r["vendor"]).font = Font(name=FONT)
        ws.cell(row, 3, r["src"]).font = Font(name=FONT)
        ws.cell(row, 4, r["dst"]).font = Font(name=FONT)
        ws.cell(row, 5, f"=C{row}+D{row}").font = Font(name=FONT, bold=True)  # total = fórmula
        ws.cell(row, 6, "Sí" if r["multicast"] else "No").font = Font(name=FONT)
        ws.cell(row, 7, "Sí" if r["local"] else "No").font = Font(name=FONT)
        for c in range(1, ncols + 1):
            ws.cell(row, c).border = BORDE
        for c in (3, 4, 5):
            ws.cell(row, c).number_format = "#,##0"
        for c in (1, 6, 7):
            ws.cell(row, c).alignment = Alignment(horizontal="center")

    # Anchos
    for c, w in zip(range(1, ncols + 1), (20, 30, 12, 13, 12, 11, 17)):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{last}"

    # Formato condicional
    rng_total = f"E3:E{last}"
    ws.conditional_formatting.add(rng_total, ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="9CC3E5",
        end_type="max", end_color="2E75B6"))

    rango = f"A3:{get_column_letter(ncols)}{last}"
    # Fabricante desconocido -> naranja
    ws.conditional_formatting.add(rango, FormulaRule(
        formula=['$B3="Unknown"'], fill=PatternFill("solid", fgColor=NARANJA)))
    # Multicast -> gris
    ws.conditional_formatting.add(rango, FormulaRule(
        formula=['$F3="Sí"'], fill=PatternFill("solid", fgColor=GRIS)))
    # MAC aleatorizada -> resalta esa celda
    ws.conditional_formatting.add(f"G3:G{last}", FormulaRule(
        formula=['$G3="Sí"'], fill=PatternFill("solid", fgColor=AMAR)))

    return last


def hoja_fabricantes(wb, filas, last_inv):
    ws = wb.create_sheet("Por fabricante")
    # Vendors únicos ordenados por nº de dispositivos (solo para ordenar las filas)
    conteo = {}
    for r in filas:
        conteo[r["vendor"]] = conteo.get(r["vendor"], 0) + 1
    vendors = sorted(conteo, key=lambda v: conteo[v], reverse=True)

    cab = ["Fabricante", "Dispositivos", "Paquetes totales"]
    for c, txt in enumerate(cab, start=1):
        cell = ws.cell(1, c, txt)
        cell.font = Font(name=FONT, bold=True, color=BLANCO)
        cell.fill = PatternFill("solid", fgColor=AZUL2)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDE

    inv = "Inventario"
    bcol = f"'{inv}'!$B$3:$B$%d" % last_inv
    ecol = f"'{inv}'!$E$3:$E$%d" % last_inv
    for i, v in enumerate(vendors):
        row = 2 + i
        ws.cell(row, 1, v).font = Font(name=FONT)
        # Conteos como fórmulas, así se recalculan si editás el inventario
        ws.cell(row, 2, f'=COUNTIF({bcol},A{row})').number_format = "#,##0"
        ws.cell(row, 3, f'=SUMIF({bcol},A{row},{ecol})').number_format = "#,##0"
        for c in range(1, 4):
            ws.cell(row, c).border = BORDE
            ws.cell(row, c).font = Font(name=FONT)

    total_row = 2 + len(vendors)
    ws.cell(total_row, 1, "Total").font = Font(name=FONT, bold=True)
    ws.cell(total_row, 2, f"=SUM(B2:B{total_row-1})").font = Font(name=FONT, bold=True)
    ws.cell(total_row, 3, f"=SUM(C2:C{total_row-1})").font = Font(name=FONT, bold=True)
    for c in range(1, 4):
        ws.cell(total_row, c).fill = PatternFill("solid", fgColor=GRIS)
        ws.cell(total_row, c).border = BORDE
        ws.cell(total_row, c).number_format = "#,##0"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 17
    ws.freeze_panes = "A2"

    # Gráfico de barras: dispositivos por fabricante (hasta 15)
    top = min(len(vendors), 15)
    chart = BarChart()
    chart.title = "Dispositivos por fabricante"
    chart.type = "bar"
    chart.legend = None
    data = Reference(ws, min_col=2, min_row=1, max_row=1 + top)
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + top)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = max(6, top * 0.6)
    chart.width = 16
    ws.add_chart(chart, "E2")


def hoja_resumen(wb, last_inv):
    ws = wb.create_sheet("Resumen", 0)  # primera hoja
    inv = "Inventario"
    b = f"'{inv}'!$B$3:$B$%d" % last_inv
    e = f"'{inv}'!$E$3:$E$%d" % last_inv
    f = f"'{inv}'!$F$3:$F$%d" % last_inv
    g = f"'{inv}'!$G$3:$G$%d" % last_inv

    ws.merge_cells("A1:C1")
    t = ws.cell(1, 1, "Resumen del inventario")
    t.font = Font(name=FONT, bold=True, size=14, color=BLANCO)
    t.fill = PatternFill("solid", fgColor=AZUL)
    ws.row_dimensions[1].height = 24

    metricas = [
        ("Dispositivos únicos", f"=COUNTA({b})"),
        ("Fabricantes identificados", f'=COUNTA({b})-COUNTIF({b},"Unknown")'),
        ("Desconocidos (Unknown)", f'=COUNTIF({b},"Unknown")'),
        ("Multicast / broadcast", f'=COUNTIF({f},"Sí")'),
        ("MAC aleatorizadas", f'=COUNTIF({g},"Sí")'),
        ("Paquetes totales", f"=SUM({e})"),
    ]
    for i, (k, formula) in enumerate(metricas):
        row = 3 + i
        kc = ws.cell(row, 1, k)
        kc.font = Font(name=FONT, bold=True)
        kc.fill = PatternFill("solid", fgColor="EDEDED")
        kc.border = BORDE
        vc = ws.cell(row, 2, formula)
        vc.font = Font(name=FONT)
        vc.number_format = "#,##0"
        vc.border = BORDE

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    # Identificados vs desconocidos para un gráfico de torta
    ws.cell(11, 1, "Identificados").font = Font(name=FONT)
    ws.cell(11, 2, f'=COUNTA({b})-COUNTIF({b},"Unknown")')
    ws.cell(12, 1, "Desconocidos").font = Font(name=FONT)
    ws.cell(12, 2, f'=COUNTIF({b},"Unknown")')
    pie = PieChart()
    pie.title = "Identificados vs desconocidos"
    data = Reference(ws, min_col=2, min_row=11, max_row=12)
    cats = Reference(ws, min_col=1, min_row=11, max_row=12)
    pie.add_data(data)
    pie.set_categories(cats)
    pie.height = 7
    pie.width = 10
    ws.add_chart(pie, "D3")

    nota = ws.cell(14, 1, f"Generado {datetime.now():%Y-%m-%d %H:%M} a partir del inventario de network_inventory")
    nota.font = Font(name=FONT, italic=True, size=9, color="808080")


def main():
    ap = argparse.ArgumentParser(description="Inventario de red -> Excel con formato")
    ap.add_argument("entrada", help="archivo CSV o JSON del inventario")
    ap.add_argument("-o", "--output", help="archivo .xlsx de salida")
    args = ap.parse_args()

    out = args.output or os.path.splitext(args.entrada)[0] + ".xlsx"
    filas = leer_inventario(args.entrada)
    if not filas:
        raise SystemExit("El inventario está vacío.")

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True  # que Excel/LibreOffice recalculen al abrir
    last = hoja_inventario(wb, filas)
    hoja_fabricantes(wb, filas, last)
    hoja_resumen(wb, last)
    wb.save(out)
    print(f"OK: {len(filas)} dispositivos -> {out}")


if __name__ == "__main__":
    main()
